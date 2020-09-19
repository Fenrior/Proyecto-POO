# Modulo para crear archivos de excel y manipular la informacion de la base de datos
# Agregado el 19/09/2020 Fernando Lavarreda
import os
import socket
import gspread as gs
import openpyxl as xl
import matplotlib.pyplot as plt

class ManejoInfo:
    def __init__(self, ruta_json):
        self.route = ruta_json
        self.downloaded = False
        self.vals1 = []
        self.vals2 = []

    def descargar_info(self):
        if self.check_connection():
            try:
                client = gs.service_account(self.route)
                sheet = client.open("Mindfulness").sheet1
            except Exception as e:
                with open("resources/error.txt", "w") as error:
                    error.write(e)
            else:
                self.vals1 = sheet.col_values(1)
                self.vals2 = sheet.col_values(2)
                self.downloaded = True
                
        
    def crear_excel(self, ubicacion, nombre):
        wb = xl.Workbook()
        wksheet = wb.active
        for value in range(len(self.vals1)):
            wksheet.cell(row=value+1, column=1).value = self.vals1[value]
            wksheet.cell(row=value+1, column=2).value = self.vals2[value]
        wb.save(ubicacion+nombre+".xlsx")
        os.startfile(ubicacion+nombre+".xlsx")

    def bar_graph(self):
        all_diseases = {}
        for value in self.vals1:
            if "|" in value:
                diseases = value.split("|")
                for d in diseases:
                    if d not in all_diseases.keys():
                        all_diseases[d] = 1
                    else:
                        all_diseases[d] += 1
            else:
                if value not in all_diseases.keys():
                    all_diseases[value] = 1
                else:
                    all_diseases[value] += 1
        plt.bar(all_diseases.keys(), all_diseases.values())
        plt.title("Diagnosticos Realizados")
        plt.show()

    
    def check_connection(self):
        """Vericar conneccion a Internet, retorna bool"""
        try:
            socket.getaddrinfo("google.com", "321")
        except socket.gaierror:
            connection = False
        else:
            connection = True
        return connection

if __name__ == "__main__":
    test = ManejoInfo("resources/creds.json")
    test.descargar_info()
    test.crear_excel("", "hola")
    test.bar_graph()